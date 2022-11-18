from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import tkinter as tk
from tkinter import filedialog
import warnings



#Getting File Path
root = tk.Tk()
root.withdraw()

fie_path = filedialog.askopenfilename()
print(fie_path)


###Opening workbook
with warnings.catch_warnings(record=True):
    warnings.simplefilter("always")
    wb = load_workbook(fie_path)

ws = wb['Sheet1']


    
#######Working with data#####
num_rows = ws.max_row
total_row_list = []

total_2016 = []
total_2017 = []
total_2018 = []
total_2019 = []
total_2020 = []
total_2021 = []

summary = dict()



###Delete Unnecessary Columns
ws.delete_cols(7,20)
ws.delete_cols(3,1)

###Get Coverage Names
coverage_names = []
for row in range(2,num_rows+1):
    coverage_names.append(ws['A' + str(row)].value)

coverage_names = set(coverage_names)    

###Every Cell
for row in range(2,num_rows + 1):
    row_list = []
    for col in range(1,6):
        char = get_column_letter(col)
        row_list.append(ws[char + str(row)].value)
        
    date = row_list[4]
    row_list[4] = str(date.year)
    
    if row_list[4] == '2016':
        total_2016.append(row_list)
    elif row_list[4] == '2017':
        total_2017.append(row_list)
    elif row_list[4] == '2018':
        total_2018.append(row_list)
    elif row_list[4] == '2019':
        total_2019.append(row_list)
    elif row_list[4] == '2020':
        total_2020.append(row_list)
    elif row_list[4] == '2021':
        total_2021.append(row_list)

###Count Duplicate Coverage Types in Total Lists
coverages_2016 = dict()
coverages_2017 = dict()
coverages_2018 = dict()
coverages_2019 = dict()
coverages_2020 = dict()
coverages_2021 = dict()
for i in range(0, len(total_2016)):
    if total_2016[i][0] in coverages_2016:
        coverages_2016[total_2016[i][0]] += 1
    else:
        coverages_2016[total_2016[i][0]] = 1        
for i in range(0, len(total_2017)):
    if total_2017[i][0] in coverages_2017:
        coverages_2017[total_2017[i][0]] += 1
    else:
        coverages_2017[total_2017[i][0]] = 1        
for i in range(0, len(total_2018)):
    if total_2018[i][0] in coverages_2018:
        coverages_2018[total_2018[i][0]] += 1
    else:
        coverages_2018[total_2018[i][0]] = 1        
for i in range(0, len(total_2019)):
    if total_2019[i][0] in coverages_2019:
        coverages_2019[total_2019[i][0]] += 1
    else:
        coverages_2019[total_2019[i][0]] = 1        
for i in range(0, len(total_2020)):
    if total_2020[i][0] in coverages_2020:
        coverages_2020[total_2020[i][0]] += 1
    else:
        coverages_2020[total_2020[i][0]] = 1        
for i in range(0, len(total_2021)):
    if total_2021[i][0] in coverages_2021:
        coverages_2021[total_2021[i][0]] += 1
    else:
        coverages_2021[total_2021[i][0]] = 1
    

###Find max num coverages
coverage_counts = dict()
for i in coverage_names:
    coverage_counts[i] = 0
    
for i in coverages_2016:
    if coverages_2016[i] > coverage_counts[i]:
        coverage_counts[i] = coverages_2016[i]
for i in coverages_2017:
    if coverages_2017[i] > coverage_counts[i]:
        coverage_counts[i] = coverages_2017[i]
for i in coverages_2018:
    if coverages_2018[i] > coverage_counts[i]:
        coverage_counts[i] = coverages_2018[i]
for i in coverages_2019:
    if coverages_2019[i] > coverage_counts[i]:
        coverage_counts[i] = coverages_2019[i]
for i in coverages_2020:
    if coverages_2020[i] > coverage_counts[i]:
        coverage_counts[i] = coverages_2020[i]
for i in coverages_2021:
    if coverages_2021[i] > coverage_counts[i]:
        coverage_counts[i] = coverages_2021[i]
        
#####Begin Building New Sheet#####    

wb.create_sheet("Program Summary")
ws = wb["Program Summary"]

###Format###
columns_to_fill = ['C', 'D', 'F', 'G', 'I', 'J', 'L', 'M', 'O', 'P', 'R', 'S']

ws.merge_cells('C1:D1')
ws.merge_cells('F1:G1')
ws.merge_cells('I1:J1')
ws.merge_cells('L1:M1')
ws.merge_cells('O1:P1')
ws.merge_cells('R1:S1')

for i in columns_to_fill:
    ws[i + '1'].alignment = Alignment(horizontal='center')


blackFill = PatternFill(start_color='00000000', end_color='00000000',
                        fill_type='solid')
purpleFill = PatternFill(start_color='00800080', end_color='00800080',
                        fill_type='solid')
yellowFill = PatternFill(start_color='00FFFF00', end_color='00FFFF00',
                        fill_type='solid')
redFill = PatternFill(start_color='00FF0000', end_color='00FF0000',
                        fill_type='solid')
greenFill = PatternFill(start_color='00008000', end_color='00008000',
                        fill_type='solid')

columnFill = PatternFill(start_color='00000000', end_color='00969696',
                         fill_type='darkGrid')
cellBorder = Border(left=Side(border_style='thin',color='00000000'),
                    right=Side(border_style='thin',color='00000000'),
                    top=Side(border_style='thin',color='00000000'),
                    bottom=Side(border_style='thin',color='00000000'))

#Legend
ws['A1'] = 'Legend'
ws['A1'].font = Font(bold=True)
ws['A2'] = 'No Policy in Place'
ws['A2'].fill = blackFill
ws['A2'].font = Font(color='00FFFFFF')
ws['A3'] = 'No Claims - Disregard Loss Run'
ws['A3'].fill = purpleFill
ws['A3'].font = Font(color='00FFFFFF')
ws['A4'] = 'Loss Run Needed'
ws['A4'].fill = yellowFill
ws['A5'] = 'Loss Run Not Available'
ws['A5'].fill = redFill
ws['A6'] = 'Loss Run Received'
ws['A6'].fill = greenFill

#Header
ws['C1'] = '1/2016 to 1/2017'
ws['C1'].font = Font(color='00339966')
ws['F1'] = '1/2017 to 1/2018'
ws['F1'].font = Font(color='00339966')
ws['I1'] = '1/2018 to 1/2019'
ws['I1'].font = Font(color='00339966')
ws['L1'] = '1/2019 to 1/2020'
ws['L1'].font = Font(color='00339966')
ws['O1'] = '1/2020 to 1/2021'
ws['O1'].font = Font(color='00339966')
ws['R1'] = '1/2021 to 1/2022'
ws['R1'].font = Font(color='00339966')

ws['B2'] = 'Coverage Type'
ws['B2'].font = Font(bold=True)

ws['C2'] = 'Policy Number'
ws['C2'].font = Font(bold=True)
ws['D2'] = 'Carrier'
ws['D2'].font = Font(bold=True)
ws['F2'] = 'Policy Number'
ws['F2'].font = Font(bold=True)
ws['G2'] = 'Carrier'
ws['G2'].font = Font(bold=True)
ws['I2'] = 'Policy Number'
ws['I2'].font = Font(bold=True)
ws['J2'] = 'Carrier'
ws['J2'].font = Font(bold=True)
ws['L2'] = 'Policy Number'
ws['L2'].font = Font(bold=True)
ws['M2'] = 'Carrier'
ws['M2'].font = Font(bold=True)
ws['O2'] = 'Policy Number'
ws['O2'].font = Font(bold=True)
ws['P2'] = 'Carrier'
ws['P2'].font = Font(bold=True)
ws['R2'] = 'Policy Number'
ws['R2'].font = Font(bold=True)
ws['S2'] = 'Carrier'
ws['S2'].font = Font(bold=True)

#Column Width
ws.column_dimensions['E'].width = 5
ws.column_dimensions['H'].width = 5
ws.column_dimensions['K'].width = 5
ws.column_dimensions['N'].width = 5
ws.column_dimensions['Q'].width = 5
ws.column_dimensions['T'].width = 5

ws.column_dimensions['A'].width = 29.71
ws.column_dimensions['B'].width = 30

ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 18
ws.column_dimensions['I'].width = 18
ws.column_dimensions['J'].width = 18
ws.column_dimensions['L'].width = 18
ws.column_dimensions['M'].width = 18
ws.column_dimensions['O'].width = 18
ws.column_dimensions['P'].width = 18
ws.column_dimensions['R'].width = 18
ws.column_dimensions['S'].width = 18

###Data input###
final_coverages = []

for i in coverage_counts:
    for n in range(coverage_counts[i]):
        if n == 0:
            final_coverages.append(i)
        else:
            final_coverages.append(i + ' ' + str(n+1))
        
for row in range(len(final_coverages)):
    ws['B'+str(row+3)] = final_coverages[row]


##Renaming group lists to math coverage names/duplicates
final_cov_2016 = []
final_cov_2017 = []
final_cov_2018 = []
final_cov_2019 = []
final_cov_2020 = []
final_cov_2021 = []

for n in range(len(total_2016)):
    k = 2
    if total_2016[n][0] not in final_cov_2016:
        final_cov_2016.append(total_2016[n][0])
    else:
        while total_2016[n][0] + " " + str(k) in final_cov_2016:
            k+=1
        final_cov_2016.append(total_2016[n][0] + " " + str(k))
        total_2016[n][0] = total_2016[n][0] + " " + str(k)
for n in range(len(total_2017)):
    k = 2
    if total_2017[n][0] not in final_cov_2017:
        final_cov_2017.append(total_2017[n][0])
    else:
        while total_2017[n][0] + " " + str(k) in final_cov_2017:
            k+=1
        final_cov_2017.append(total_2017[n][0] + " " + str(k))
        total_2017[n][0] = total_2017[n][0] + " " + str(k)
for n in range(len(total_2018)):
    k = 2
    if total_2018[n][0] not in final_cov_2018:
        final_cov_2018.append(total_2018[n][0])
    else:
        while total_2018[n][0] + " " + str(k) in final_cov_2018:
            k+=1
        final_cov_2018.append(total_2018[n][0] + " " + str(k))
        total_2018[n][0] = total_2018[n][0] + " " + str(k)
for n in range(len(total_2019)):
    k = 2
    if total_2019[n][0] not in final_cov_2019:
        final_cov_2019.append(total_2019[n][0])
    else:
        while total_2019[n][0] + " " + str(k) in final_cov_2019:
            k+=1
        final_cov_2019.append(total_2019[n][0] + " " + str(k))
        total_2019[n][0] = total_2019[n][0] + " " + str(k)

for n in range(len(total_2020)):
    k = 2
    if total_2020[n][0] not in final_cov_2020:
        final_cov_2020.append(total_2020[n][0])
    else:
        while total_2020[n][0] + " " + str(k) in final_cov_2020:
            k+=1
        final_cov_2020.append(total_2020[n][0] + " " + str(k))
        total_2020[n][0] = total_2020[n][0] + " " + str(k)
for n in range(len(total_2021)):
    k = 2
    if total_2021[n][0] not in final_cov_2021:
        final_cov_2021.append(total_2021[n][0])
    else:
        while total_2021[n][0] + " " + str(k) in final_cov_2021:
            k+=1
        final_cov_2021.append(total_2021[n][0] + " " + str(k))
        total_2021[n][0] = total_2021[n][0] + " " + str(k)

#Matching up data with coverage types
for i in range(3,len(final_coverages) + 3):
    for n in total_2016:
        if n[0] == ws['B'+str(i)].value:
            ws['C'+str(i)] = n[1]
            ws['D'+str(i)] = n[3]
    for n in total_2017:
        if n[0] == ws['B'+str(i)].value:
            ws['F'+str(i)] = n[1]
            ws['G'+str(i)] = n[3]
    for n in total_2018:
        if n[0] == ws['B'+str(i)].value:
            ws['I'+str(i)] = n[1]
            ws['J'+str(i)] = n[3]
    for n in total_2019:
        if n[0] == ws['B'+str(i)].value:
            ws['L'+str(i)] = n[1]
            ws['M'+str(i)] = n[3]
    for n in total_2020:
        if n[0] == ws['B'+str(i)].value:
            ws['O'+str(i)] = n[1]
            ws['P'+str(i)] = n[3]
    for n in total_2021:
        if n[0] == ws['B'+str(i)].value:
            ws['R'+str(i)] = n[1]
            ws['S'+str(i)] = n[3]

#Filling Columns Between Years
for row in range(1, len(final_coverages) + 3):
    ws['E'+str(row)].fill = columnFill
    ws['H'+str(row)].fill = columnFill
    ws['K'+str(row)].fill = columnFill
    ws['N'+str(row)].fill = columnFill
    ws['Q'+str(row)].fill = columnFill
    ws['T'+str(row)].fill = columnFill

#Filling color in policy number and carrier
for i in range(3, len(final_coverages) + 3):
    for n in columns_to_fill:
        if ws[n + str(i)].value == None:
            ws[n + str(i)].fill = blackFill
        else:
            ws[n + str(i)].fill = yellowFill
        ws[n + str(i)].border = cellBorder

#####Saving Worksheet#####
save_name = str(input('Input a name for the file: '))
input('Now select a folder to save the file to. Press any key to continue.')
save_to = filedialog.askdirectory()

wb.save(save_to + '/' + save_name + '.xlsx')
